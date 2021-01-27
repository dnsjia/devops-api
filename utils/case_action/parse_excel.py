#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @File  : parse_excel.py
# @Author: 风哥
# @Email: gujiwork@outlook.com
# @Date  : 2021/1/10
# @Desc  :
import traceback
from openpyxl import load_workbook
from datetime import datetime
from utils.case_action.config_read import TEST_STEP_END_TIME, TEST_STEP_ERROR, TEST_STEP_PICTURE, TEST_STEP_RESULT, \
    TEST_CASE_RESULT, TEST_CASE_END_TIME
from openpyxl.drawing.image import Image
import xlrd
from auto_case.models import TestTask
import logging
from utils.case_action.case_logs import CaseInsertLogs
logger = logging.getLogger('default')


class ParseExcel(object):
    """
    解析EXCEL文档
    """

    def __init__(self, file_name, data):
        self.file_name = file_name
        self.data = data
        self.task_obj = TestTask.objects.filter(task_id=self.data.get('task_id')).first()

        # 读取excel文件
        try:
            self.wb = load_workbook(self.file_name)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, '开始读取用例文件', 3)
        except BaseException as e:
            msg = '无法读取或已损坏, 异常原因：%s' % str(traceback.format_exc())
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
            logger.error('文件：%s, 无法读取或已损坏, 异常原因：%s' % (self.file_name, str(traceback.format_exc())), e)

    def get_row_value(self, sheet_name, row_no):
        """
        获取excel某一行的数据
        :param sheet_name:
        :param row_no:
        :return:
        """
        print('get_row_value')
        try:
            # sheet_name
            sheet_names = self.wb[sheet_name]
            # 创建集合，将指定行内的数据添加进集合
            row_value_list = []
            # 循环所有列
            for i in range(1, sheet_names.max_column + 1):
                # 通过行号与列号获取指定单元格信息，并添加进集合
                value = sheet_names.cell(row_no, i).value
                row_value_list.append(value)
            return row_value_list
        except Exception as e:
            msg = '读取失败，请检查工作表名以及行，列号, 异常原因：%s' % str(traceback.format_exc())
            logger.error(msg, e)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)

        finally:
            self.wb.close()
            print('结束getRowValue')

    def get_column_value(self, sheet_name, column_no):
        """
        获取excel某一列的数据
        :param sheet_name:
        :param column_no:
        :return:
        """
        try:
            sheet_names = self.wb[sheet_name]
            column_value_list = []
            for i in range(2, sheet_names.max_row + 1):
                value = sheet_names.cell(i, column_no).value
                column_value_list.append(value)
            return column_value_list
        except Exception as e:
            msg = '读取失败，请检查工作表名以及行，列号, 异常原因：%s' % str(traceback.format_exc())
            logger.error(msg, e)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
        finally:
            self.wb.close()

    def get_merge_column_value(self, sheet_name, column_no):
        """
        读取合并单元格的数据
        :param sheet_name: 工作表
        :param column_no: 列号
        :return:
        """
        try:
            # 获取数据
            data = xlrd.open_workbook(self.file_name)
            # 获取sheet
            sheet_name = data.sheet_by_name(sheet_name)
            # 获取总行数
            n_rows = sheet_name.nrows  # 包括标题
            # 获取总列数
            n_cols = sheet_name.ncols
            # 计算出合并的单元格有哪些
            colspan = {}
            # 如果sheet是合并的单元格 则获取合并单元格的值，并将第一行的数据赋值给合并单元格中的空值
            if sheet_name.merged_cells:
                for item in sheet_name.merged_cells:
                    for row in range(item[0], item[1]):
                        for col in range(item[2], item[3]):
                            # 合并单元格的首格是有值的，所以在这里进行了去重
                            if (row, col) != (item[0], item[2]):
                                colspan.update({(row, col): (item[0], item[2])})

                col = []
                for i in range(1, n_rows):
                    if colspan.get((i, column_no - 1)):
                        value = sheet_name.cell_value(*colspan.get((i, column_no - 1)))
                        col.append(value)
                    else:
                        col.append(sheet_name.cell_value(i, column_no - 1))
                return col
        except Exception as e:
            msg = '合并单元格读取错误, 异常原因：%s' % str(traceback.format_exc())
            logger.error(msg, e)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
        finally:
            self.wb.close()

    def is_merge(self, sheet_name):
        """
        判断'工作表'内是否有合并单元格
        :param sheet_name:
        :return:
        """
        sheet_names = self.wb[sheet_name]
        merge = sheet_names.merged_cells
        return merge

    def get_cell_value(self, sheet_name, row_no, column_no):
        """
        获取excel某一单元格的数据
        :param sheet_name:
        :param row_no:
        :param column_no:
        :return:
        """
        try:
            sheet_names = self.wb[sheet_name]
            cell_value = sheet_names.cell(row_no, column_no).value
            return cell_value
        except Exception as e:
            msg = '读取失败，请检查工作表名以及行，列号, 异常原因：%s' % str(traceback.format_exc())
            logger.error(msg, e)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
        finally:
            self.wb.close()

    def get_cell_object(self, sheet_name, row_no, column_no):
        """
        获取excel某一单元格的数据
        :param sheet_name:
        :param row_no:
        :param column_no:
        :return:
        """
        try:
            sheet_names = self.wb[sheet_name]
            cell_value = sheet_names.cell(row_no, column_no)
            return cell_value
        except Exception as e:
            msg = '读取失败，请检查工作表名以及行，列号, 异常原因：%s' % str(traceback.format_exc())
            logger.error(msg, e)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
        finally:
            self.wb.close()

    def write_cell_value(self, sheet_name, row_no, column_no, value):
        """
        向excel某一单元格写入数据
        :param sheet_name:
        :param row_no:
        :param column_no:
        :param value:
        :return:
        """
        try:
            sheet_names = self.wb[sheet_name]
            sheet_names.cell(row_no, column_no, value)
            self.wb.save(self.file_name)
        except PermissionError:
            msg = '请先关闭用例文件，再运行测试用例'
            logger.warning(msg)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
            raise
        except Exception as e:
            msg = '写入失败，请检查工作表名以及行，列号, %s' % str(traceback.format_exc())
            logger.error(msg, e)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
        finally:
            self.wb.close()

    def write_cell_time(self, sheet_name, row_no, column_no):
        """
        向excel某一单元格写入时间
        :param sheet_name:
        :param row_no:
        :param column_no:
        :return:
        """
        try:
            sh = self.wb[sheet_name]
            now_time = datetime.now()
            now_time.strftime('%Y:%m:%d %H:%M:%S')
            sh.cell(row_no, column_no, now_time)
            self.wb.save(self.file_name)
            logger.info('%s写入时间成功，写入时间为：%s' % (self.file_name, now_time))
        except PermissionError:
            msg = '请先关闭用例文件，再运行测试用例'
            logger.warning(msg)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
        except Exception as e:
            msg = '写入失败，请检查工作表名以及行，列号, 异常原因：%s' % str(traceback.format_exc())
            logger.error(msg, e)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
        finally:
            self.wb.close()

    def write_cell_values(self, sheet_name, row_no, err_info=None, err_pic=None):
        """
        写入 错误信息 错误图片
        :param sheet_name:
        :param row_no:
        :param err_info:
        :param err_pic:
        :return:
        """
        try:
            self.write_cell_time(sheet_name, row_no, TEST_STEP_END_TIME)
            if err_info:
                self.write_cell_value(sheet_name, row_no, TEST_STEP_ERROR, err_info)
            elif err_pic:
                self.write_cell_value(sheet_name, row_no, TEST_STEP_PICTURE, err_pic)
            else:
                self.write_cell_value(sheet_name, row_no, TEST_STEP_ERROR, '')
                self.write_cell_value(sheet_name, row_no, TEST_STEP_PICTURE, '')
            self.wb.save(self.file_name)
            logger.info('用例测试结果，错误信息，错误图片写入成功')
        except PermissionError:
            msg = '请先关闭用例文件，再运行测试用例'
            logger.warning(msg)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
        except Exception as e:
            self.wb.close()
            msg = '用例测试结果，错误信息，错误图片写入失败, 异常原因：%s' % str(traceback.format_exc())
            logger.error(msg, e)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
        finally:
            self.wb.close()

    def clear_cell_value(self, sheet_name, row_no):
        """
        清空EXCEL单元格数据
        :param sheet_name:
        :param row_no:
        :return:
        """
        try:
            end_time = self.get_cell_value(sheet_name, row_no, TEST_STEP_END_TIME)
            result = self.get_cell_value(sheet_name, row_no, TEST_STEP_RESULT)
            err_info = self.get_cell_value(sheet_name, row_no, TEST_STEP_ERROR)
            err_pic = self.get_cell_value(sheet_name, row_no, TEST_STEP_PICTURE)
            if end_time is not None or end_time != '' and result is not None or result != '' and \
                    err_info is not None or err_info != '' and err_pic is not None or err_pic != '':
                self.write_cell_value(sheet_name, row_no, TEST_STEP_END_TIME, '')
                self.write_cell_value(sheet_name, row_no, TEST_STEP_RESULT, '')
                self.write_cell_value(sheet_name, row_no, TEST_STEP_ERROR, '')
                self.write_cell_value(sheet_name, row_no, TEST_STEP_PICTURE, '')
            self.wb.save(self.file_name)
        except PermissionError:
            msg = '请先关闭用例文件，再运行测试用例'
            logger.warning(msg)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
        except Exception as e:
            msg = '数据清空失败, 原因：%s' % str(traceback.format_exc())
            logger.error(msg, e)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
        finally:
            self.wb.close()

    def clear_step_column_value(self, sheet_name):
        """
        清除执行时间，错误结果，错误信息，错误截图信息
        :param sheet_name:
        :return:
        """
        try:
            msg = '清除"%s"工作表测试结果中，请稍等...' % sheet_name
            logger.info(msg)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
            end_times = self.get_column_value(sheet_name, TEST_STEP_END_TIME)
            err_infos = self.get_column_value(sheet_name, TEST_STEP_ERROR)
            err_pics = self.get_column_value(sheet_name, TEST_STEP_PICTURE)
            for a, b in enumerate(end_times):
                if b == '测试执行时间':
                    continue
                elif b != '' or b is not None:
                    self.wb[sheet_name].cell(a + 2, TEST_STEP_END_TIME, '')
            for e, f in enumerate(err_infos):
                if f == '错误信息':
                    continue
                elif f != '' or f is not None:
                    self.wb[sheet_name].cell(e + 2, TEST_STEP_ERROR, '')
            for g, h in enumerate(err_pics):
                if h == '测试截图':
                    continue
                elif h != '' or h is not None:
                    self.wb[sheet_name].cell(g + 2, TEST_STEP_PICTURE, '')
            # 清除用例的测试结果
            for l in range(5):
                results = self.get_column_value(sheet_name, TEST_STEP_RESULT + l)
                none_results = list(filter(None, self.get_column_value(sheet_name, TEST_STEP_RESULT + l)))
                if len(none_results) == 0:
                    continue
                else:
                    for c, d in enumerate(results):
                        if '测试结果1' == d or '测试结果2' == d or '测试结果3' == d or '测试结果4' == d or '测试结果5' == d:
                            continue
                        elif type(
                                self.get_cell_object(sheet_name, c + 2, TEST_STEP_RESULT + l)).__name__ == 'MergedCell':
                            continue
                        elif d != '' or d is not None:
                            self.wb[sheet_name].cell(c + 2, TEST_STEP_RESULT + l, '')
            self.wb.save(self.file_name)
        except PermissionError:
            msg = '请先关闭用例文件，再运行测试用例'
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
        except Exception as e:
            msg = '数据清空失败, 异常原因：%s' % str(traceback.format_exc())
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
            logger.error(msg, e)
        finally:
            self.wb.close()

    def clear_case_column_value(self, sheet_name):
        """
        清除执行时间，错误结果，错误信息，错误截图信息
        :param sheet_name:
        :return:
        """
        try:
            msg = '清除"%s"工作表测试结果中，请稍等....' % sheet_name
            logger.info(msg)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
            # 清除用例的测试结果
            for l in range(5):
                case_result = self.get_column_value(sheet_name, TEST_CASE_RESULT + l)
                none_case_result = list(filter(None, self.get_column_value(sheet_name, TEST_CASE_RESULT + l)))
                if len(none_case_result) == 1:
                    continue
                else:
                    for i, v in enumerate(case_result):
                        if '执行结果1' == v or '执行结果2' == v or '执行结果3' == v or '执行结果4' == v or '执行结果5' == v:
                            continue
                        elif type(
                                self.get_cell_object(sheet_name, i + 2, TEST_CASE_RESULT + l)).__name__ == 'MergedCell':
                            continue
                        elif v != '' or v is not None:
                            self.wb[sheet_name].cell(i + 2, TEST_CASE_RESULT + l, '')
            # 清除执行时间
            case_time = self.get_column_value(sheet_name, TEST_CASE_END_TIME)
            for s, d in enumerate(case_time):
                if d == '执行结束时间':
                    continue
                elif type(self.get_cell_object(sheet_name, s + 2, TEST_CASE_END_TIME)).__name__ == 'MergedCell':
                    continue
                elif d != '' or d is not None:
                    self.wb[sheet_name].cell(s + 2, TEST_CASE_END_TIME, '')
            self.wb.save(self.file_name)
        except PermissionError:
            msg = '请先关闭用例文件，再运行测试用例'
            logger.warning(msg)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
        except Exception as e:
            msg = '数据清空失败, 异常原因：%s' % str(traceback.format_exc())
            logger.error(msg, e)
            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
        finally:
            self.wb.close()
