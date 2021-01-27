#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @File  : testcase.py
# @Author: 风哥
# @Email: gujiwork@outlook.com
# @Date  : 2021/1/10
# @Desc  :
from django_redis import get_redis_connection
from utils.case_action.case_logs import CaseInsertLogs
from utils.case_action.parse_excel import ParseExcel
from utils.case_action.config_read import EXCEL_TEMPLATE_PATH, TEST_CASE_IS_IMPL_EMENT, TEST_CASE_SHEET, TEST_CASE_NUM, \
    TEST_STEP_NUM, TEST_STEP_PRE_SET, BROWSER_VERSION, TEST_STEP_DESCRIBE, TEST_STEP_KEYWORD, TEST_STEP_LOCATION, \
    TEST_STEP_LOCATOR, TEST_STEP_VALUE, TEST_STEP_RESULT, TEST_STEP_END_TIME, TEST_STEP_ERROR, TEST_STEP_PICTURE, \
    TEST_CASE_END_TIME, TEST_CASE_RESULT, REPORT_PATH
from utils.case_action.page_action import PageAction
from selenium.common.exceptions import *  # 导入所有异常类
from utils.test_case.HTMLTestRunner_cn import HTMLTestRunner
from datetime import datetime
from openpyxl.styles import Font
from BeautifulReport import BeautifulReport as bf
from utils.case_action.dir_and_time import DirAndTime
from utils.case_action.case_notice import CaseDoneNotice
import traceback
import time
import re
import unittest
import logging
from auto_case.models import TestTask, TestCaseDetail
conn_redis = get_redis_connection('default')
logger = logging.getLogger('default')


class RunnerTestCase(unittest.TestCase):
    """ TestCase classes that want to be parametrized should
     inherit from this class.
    """
    def __init__(self, methodName, data):
        super(RunnerTestCase, self).__init__(methodName)
        self.data = data
        self.task_obj = TestTask.objects.filter(task_id=self.data.get('task_id')).first()
        self.start_time = datetime.now()
        self.LOCK_CASE_ID = data['case_id'] + "_lock"

    def setUp(self):
        logger.info('执行setup' + DirAndTime.get_current_time())
        logger.info('开始进来了', self.data)
        self.test_data_path = str(EXCEL_TEMPLATE_PATH) + str(self.data.get('import_address'))
        logger.info(self.test_data_path)
        self.parse_excel = ParseExcel(self.test_data_path, self.data)
        self.page_action = PageAction(self.data)
        self.sheet_names = self.parse_excel.wb.sheetnames
        self.case_num = 0
        # 创建六个字典，分别储存步骤测试结果，用例测试结果，用例测试时间，错误信息，截图信息，步骤测试时间
        self.time_dic = {}
        self.result_dic = {}
        self.error_dic = {}
        self.picture_dic = {}
        self.case_result_dic = {}
        self.case_time_dic = {}
        self.font = Font(color=None)

    def tearDown(self):
        """
        计算用例任务执行耗时，写入到Task任务表
        """
        end_time = datetime.now()
        total_time = end_time - self.start_time
        task_obj = TestTask.objects.get(task_id=self.data.get('task_id'))
        task_obj.total_time = total_time
        task_obj.save()

        success = TestCaseDetail.objects.filter(task_id=self.data.get('task_id'), status=0).count()
        failed = TestCaseDetail.objects.filter(task_id=self.data.get('task_id'), status=1).count()
        if failed > 2:
            msg = "自动化测试用例已跑完，用例失败数过多请及时关注。\n成功：%s, 失败：%s, \n任务id: %s \n 请前往运维平台查看" % (success, failed, str(self.data.get('task_id')))
            CaseDoneNotice.notice(msg)
        else:
            msg = "自动化测试用例已跑完，成功：%s, 失败：%s, 任务id: %s \n 请前往运维平台查看" % (success, failed, str(self.data.get('task_id')))
            CaseDoneNotice.notice(msg)
        # 释放任务锁
        conn_redis.delete(self.LOCK_CASE_ID)

    def test_case(self):
        print('执行test_case' + DirAndTime.get_current_time())
        try:
            # self.setUp()
            # 获取循环次数
            loop = int(self.data.get('loop', 1))
            # 获取模块名
            run_models = self.data.get('run_models', '全部')
            # 清除用例旧数据
            self.parse_excel.clear_case_column_value(self.sheet_names[0])
            # 清除步骤旧数据
            for i, v in enumerate(self.sheet_names):
                if i == 0:
                    continue
                else:
                    self.parse_excel.clear_step_column_value(v)
            for l in range(loop):
                # 用例运行数
                try:
                    # 获取'是否执行'列
                    is_impl_ement = self.parse_excel.get_column_value(self.sheet_names[0], TEST_CASE_IS_IMPL_EMENT)
                    # 循环'是否执行'列
                    # 如果执行，且模块名符合，则获取用例编号，并切换到对应的工作表，执行用例
                    for index, value in enumerate(is_impl_ement):
                        if run_models == '全部':
                            pd = "value.lower() == 'y'"
                        else:
                            pd = 'value.lower() == "y" and run_models ' \
                                 '== self.parse_excel.get_cell_value(self.sheet_names[0], index + 2, test_case_sheet)'
                        try:
                            # 如果是否执行为空则跳过执行
                            if value is None or value == '':
                                continue
                            elif eval(pd):
                                # 根据'是否执行'；列索引获取对应的工作表名
                                sheet_name = self.parse_excel.get_cell_value(self.sheet_names[0], index + 2,
                                                                             TEST_CASE_SHEET)
                                # 根据'是否执行'列索引获取对应的用例编号
                                test_case_num = self.parse_excel.get_cell_value(self.sheet_names[0], index + 2,
                                                                                TEST_CASE_NUM)
                                # 切换到用例对应的工作表
                                # sheet_names = self.parse_excel.wb[sheet_name]
                                '''
                                根据用例编号(test_case_num)获取预置条件编号
                                '''
                                # 获取用例步骤的用例编号类，并执行对应用例编号的用例步骤（增加表内是否有合并单元格的判断）
                                if self.parse_excel.is_merge(sheet_name):
                                    test_step_num = self.parse_excel.get_merge_column_value(sheet_name, TEST_STEP_NUM)
                                else:
                                    test_step_num = self.parse_excel.get_column_value(sheet_name, TEST_STEP_NUM)
                                # 循环用例步骤编号，根据索引获取预置条件编号
                                test_pre_num = ''
                                for i, v in enumerate(test_step_num):
                                    if v == test_case_num:
                                        # 用例前置条件编号
                                        test_pre_num = self.parse_excel.get_cell_value(sheet_name, i + 2,
                                                                                       TEST_STEP_PRE_SET)
                                        break

                                # 循环用例步骤编号，找到与预置条件编号相同的用例步骤编号
                                # 循环所有的步骤编号
                                # 获取对应用例编号的步骤编号的关键字，定位方式，表达式，操作值

                                url = self.data.get('ip')
                                while re.match(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}$", url) is None and re.match(r'[^\s]*[.com|.cn]', url) is None:
                                    # 从输入框获取浏览器地址
                                    url = self.data.get('ip')
                                    # 先打开浏览器，进入指定IP地址
                                    time.sleep(1)
                                self.page_action.open_browser(self.data.get('browser'), self.data.get('version', BROWSER_VERSION))
                                self.page_action.get_url(url)
                                # 执行预置条件
                                for t, v in enumerate(test_step_num):
                                    if v == test_pre_num:
                                        # 用例执行步骤
                                        pre_step_name = self.parse_excel.get_cell_value(sheet_name, t + 2, TEST_STEP_DESCRIBE)
                                        # 获取预置条件关键字
                                        pre_keyword = self.parse_excel.get_cell_value(sheet_name, t + 2, TEST_STEP_KEYWORD)
                                        # 去除前后空格
                                        if pre_keyword is not None:
                                            pre_keyword = pre_keyword.strip()
                                        # 获取定位方式
                                        pre_location = self.parse_excel.get_cell_value(sheet_name, t + 2, TEST_STEP_LOCATION)
                                        # 去除前后空格
                                        if pre_location is not None:
                                            pre_location = pre_location.strip()
                                        # 获取定位表达式
                                        pre_locator = self.parse_excel.get_cell_value(sheet_name, t + 2, TEST_STEP_LOCATOR)
                                        if type(pre_locator) is int:
                                            pre_locator = str(self.parse_excel.get_cell_value(sheet_name, t + 2, TEST_STEP_LOCATOR))
                                        # 获取输入值
                                        pre_test_value = self.parse_excel.get_cell_value(sheet_name, t + 2, TEST_STEP_VALUE)
                                        # 如果输入值为 int 类型，则强转为 str 类型，用于字符串拼接
                                        if type(pre_test_value) is int:
                                            pre_test_value = str(self.parse_excel.get_cell_value(sheet_name, t + 2, TEST_STEP_VALUE))
                                        # 总共有四种情况可以正常执行，其他情况则会将用例判断为运行失败
                                        # 1.关键字，定位方式，表达式，输入值全部不为空的情况 例：send_keys
                                        # 2.关键字，输入值不为空，定位方式，表达式为空的情况 例：assert（断言）
                                        # 3.关键字，定位方式，表达式不为空，输入值为空的情况 例：click
                                        # 4.关键字不为空，定位方式，表达式，输入值为空的情况 例 get_title
                                        if pre_keyword and pre_location and pre_locator and pre_test_value:
                                            pre_fun = 'self.page_action' + '.' + pre_keyword + '(' + '"' + pre_location + '"' + ', ' + '"' + pre_locator + '"' + ', ' + '"' + \
                                                      pre_test_value + '"' + ')'
                                        elif pre_keyword and pre_test_value and pre_location is None or pre_location == '' \
                                                and pre_locator is None or pre_location == '':
                                            pre_fun = 'self.page_action' + '.' + pre_keyword + '(' + '"' + pre_test_value + '"' + ')'
                                        elif pre_keyword and pre_location and pre_locator and pre_test_value is None or pre_test_value == '':
                                            pre_fun = 'self.page_action' + '.' + pre_keyword + '(' + '"' + pre_location + '"' + ', ' + '"' + pre_locator + '"' + ')'
                                        elif pre_keyword and pre_location is None or pre_location == '' and pre_locator is None \
                                                or pre_locator == '' and pre_test_value is None or pre_test_value == '':
                                            pre_fun = 'self.page_action' + '.' + pre_keyword + '(' + ')'
                                        elif pre_keyword is None or pre_keyword == '' and pre_location is None or pre_location == '' \
                                                and pre_locator is None or pre_locator == '' and pre_test_value is None or pre_test_value == '':
                                            continue
                                        else:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheet_name, {})[t + 2] = '关键字对应参数错误'
                                            msg = '关键字对应参数错误'
                                            logger.info(msg)
                                            print(msg)
                                            print('screenshot:', DirAndTime.get_current_time(), '.png')
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 2)
                                            continue
                                        # 执行用例
                                        try:
                                            eval(pre_fun)
                                        # 抛出异常的情况，将失败结果写入excel表格中
                                        except TypeError as e:
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheet_name, {})[t + 2] = '关键字参数个数错误，请检查参数'
                                            msg = '步骤"{}"执行失败'.format(pre_step_name)
                                            logger.info(msg, e)
                                            msg2 = '关键字参数个数错误，请检查参数'
                                            logger.info(msg2)
                                            print(msg)
                                            print(msg2)
                                            print('screenshot:', DirAndTime.get_current_time(), '.png')
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 1)
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg2, 2)
                                        except TimeoutException:
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheet_name, {})[t + 2] = '元素定位超时,请检查上一步是否执行成功,或元素定位方式'
                                            msg = '步骤"{}"执行失败'.format(pre_step_name)
                                            logger.info(msg)
                                            msg2 = '元素定位超时，请检查上一步是否执行成功，或元素定位方式'
                                            logger.info(msg2)
                                            print(msg)
                                            print(msg2)
                                            print('screenshot:', DirAndTime.get_current_time(), '.png')
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 1)
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg2, 2)
                                        except TimeoutError as e:
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Failed'
                                            self.error_dic.setdefault(sheet_name, {})[t + 2] = '断言失败'
                                            msg = '步骤"{}"执行失败'.format(pre_step_name)
                                            logger.info(msg, e)
                                            print(msg)
                                            logger.info(str(traceback.format_exc()))
                                            print('screenshot:', DirAndTime.get_current_time(), '.png')
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 1)
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, '断言失败: %s' % str(traceback.format_exc()), 3)
                                        except AttributeError as e:
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheet_name, {})[t + 2] = '元素定位超时，请检查元素定位'
                                            msg = '步骤"{}"执行失败'.format(pre_step_name)
                                            logger.info(msg, e)
                                            print(msg)
                                            logger.info(str(traceback.format_exc()))
                                            print('screenshot:', DirAndTime.get_current_time(), '.png')
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 1)
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, '元素定位超时，请检查元素定位', 2)
                                        except AssertionError:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Failed'
                                            self.error_dic.setdefault(sheet_name, {})[t + 2] = '断言失败'
                                            msg = '步骤"{}"执行失败'.format(pre_step_name)
                                            logger.info(msg)
                                            print(msg)
                                            print('screenshot:', DirAndTime.get_current_time(), '.png')
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 1)
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, '断言失败: %s' % str(traceback.format_exc()), 3)
                                        except WebDriverException as e:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheet_name, {})[t + 2] = '浏览器异常，' \
                                                                                               '请检查浏览器驱动或运行过程中是否被强制关闭'
                                            msg = '步骤"{}"执行失败'.format(pre_step_name)
                                            logger.info(msg, e)
                                            msg2 = '浏览器异常，请检查浏览器驱动或运行过程中是否被强制关闭'
                                            logger.info(msg2)
                                            print(msg)
                                            print(msg2)
                                            print('screenshot:', DirAndTime.get_current_time(), '.png')
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 1)
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg2, 2)
                                        except Exception as e:
                                            error_info = traceback.format_exc()
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheet_name, {})[t + 2] = error_info
                                            msg = '步骤"{}"执行失败'.format(pre_step_name)
                                            logger.info(msg, e)
                                            print(msg)
                                            print('screenshot:', DirAndTime.get_current_time(), '.png')
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 2)
                                        else:
                                            # 写入测试结果, 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Pass'
                                            msg = '步骤"{}"执行成功'.format(pre_step_name)
                                            logger.info(msg)
                                            print(msg)
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 0)
                                        finally:
                                            # 截图
                                            if self.result_dic.setdefault(sheet_name, {})[t + 2] == 'Skip' or \
                                                    self.result_dic.setdefault(sheet_name, {})[t + 2] == 'Failed':
                                                pic = self.page_action.save_scree_shot(sheet_name, test_case_num)
                                                self.picture_dic.setdefault(sheet_name, {})[t + 2] = pic
                                                CaseInsertLogs.test_case_logs(self.task_obj.task_id, pic, 0)
                                            # 将截图信息以及测试时间存入字典中
                                            now_time = datetime.now()
                                            now_time.strftime('%Y:%m:%d %H:%M:%S')

                                            self.time_dic.setdefault(sheet_name, {})[t + 2] = now_time
                                    else:
                                        continue

                                # 将用例步骤工作表内的用例编号以字典的方式循环
                                for t, v in enumerate(test_step_num):
                                    # 用例步骤（用例编号） 与 用例列表（用例编号）相同的
                                    if v == test_case_num:
                                        # 用例执行步骤
                                        step_name = self.parse_excel.get_cell_value(sheet_name, t + 2,TEST_STEP_DESCRIBE)
                                        # 获取关键字
                                        keyword = self.parse_excel.get_cell_value(sheet_name, t + 2, TEST_STEP_KEYWORD)
                                        # 去除前后空格
                                        if keyword is not None:
                                            keyword = keyword.strip()
                                        # 获取定位方式
                                        location = self.parse_excel.get_cell_value(sheet_name, t + 2, TEST_STEP_LOCATION)
                                        # 去除前后空格
                                        if location is not None:
                                            location = location.strip()
                                        # 获取定位表达式
                                        locator = self.parse_excel.get_cell_value(sheet_name, t + 2, TEST_STEP_LOCATOR)
                                        if type(locator) is int:
                                            locator = str(self.parse_excel.get_cell_value(sheet_name, t + 2, TEST_STEP_LOCATOR))
                                        # 获取输入值
                                        test_value = self.parse_excel.get_cell_value(sheet_name, t + 2, TEST_STEP_VALUE)
                                        # 如果输入值为 int 类型，则强转为 str 类型，用于字符串拼接
                                        if test_value is not None and type(test_value) is not str:
                                            test_value = str(self.parse_excel.get_cell_value(sheet_name, t + 2, TEST_STEP_VALUE))

                                        # 进行关键字拼接
                                        # 总共有四种情况可以正常执行，其他情况则会将用例判断为运行失败
                                        # 1.关键字，定位方式，表达式，输入值全部不为空的情况 例：send_keys
                                        # 2.关键字，输入值不为空，定位方式，表达式为空的情况 例：assert（断言）
                                        # 3.关键字，定位方式，表达式不为空，输入值为空的情况 例：click
                                        # 4.关键字不为空，定位方式，表达式，输入值为空的情况 例 getTitle
                                        if keyword and location and locator and test_value:
                                            fun = 'self.page_action' + '.' + keyword + '(' + '"' + location + '"' + ', ' + '"' + locator + '"' + ', ' + '"' + \
                                                  test_value + '"' + ')'
                                        elif keyword and test_value and location is None or location == '' \
                                                and locator is None or location == '':
                                            fun = 'self.page_action' + '.' + keyword + '(' + '"' + test_value + '"' + ')'
                                        elif keyword and location and locator and test_value is None or test_value == '':
                                            fun = 'self.page_action' + '.' + keyword + '(' + '"' + location + '"' + ', ' + '"' + locator + '"' + ')'
                                        elif keyword and location is None or location == '' and locator is None \
                                                or locator == '' and test_value is None or test_value == '':
                                            fun = 'self.page_action' + '.' + keyword + '(' + ')'
                                        elif keyword is None or keyword == '' and location is None or location == '' \
                                                and locator is None or locator == '' and test_value is None or test_value == '':
                                            continue
                                        else:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheet_name, {})[t + 2] = '关键字对应参数错误'
                                            msg = '关键字对应参数错误'
                                            logger.info(msg)
                                            print(msg)
                                            print('screenshot:', DirAndTime.get_current_time(), '.png')
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 2)
                                            continue
                                        # 执行用例
                                        try:
                                            eval(fun)
                                        # 抛出异常的情况，将失败结果写入excel表格中
                                        except TypeError as e:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheet_name, {})[t + 2] = '关键字参数个数错误，请检查参数'
                                            msg = '步骤"{}"执行失败'.format(step_name)
                                            logger.info(msg)
                                            msg2 = '关键字参数个数错误，请检查参数'
                                            logger.info(msg2)
                                            print(msg)
                                            print(msg2)
                                            logger.info(str(traceback.format_exc()))
                                            print('screenshot:', DirAndTime.get_current_time(), '.png')
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 1)
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg2, 2)
                                        except TimeoutException as e:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheet_name, {})[t + 2] = '元素定位超时，' \
                                                                                               '请检查上一步是否执行成功，或元素定位方式'
                                            msg = '步骤"{}"执行失败'.format(step_name)
                                            logger.info(msg)
                                            msg2 = '元素定位超时，请检查上一步是否执行成功，或元素定位方式'
                                            logger.info(msg2)
                                            print(msg)
                                            print(msg2)
                                            logger.info(str(traceback.format_exc()))
                                            print('screenshot:', DirAndTime.get_current_time(), '.png')
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 1)
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg2, 2)
                                        except TimeoutError as e:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Failed'
                                            self.error_dic.setdefault(sheet_name, {})[t + 2] = '断言失败'
                                            msg = '步骤"{}"执行失败'.format(step_name)
                                            logger.info(msg)
                                            msg2 = '步骤"{}"执行失败'.format(step_name)
                                            print(msg2)
                                            logger.info(str(traceback.format_exc()))
                                            print('screenshot:', DirAndTime.get_current_time(), '.png')
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 1)
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, '断言失败: %s' % str(traceback.format_exc()), 2)
                                        except AttributeError as e:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheet_name, {})[t + 2] = '元素定位超时，请检查元素定位'
                                            msg = '步骤"{}"执行失败'.format(step_name)
                                            logger.info(msg)
                                            print(msg)
                                            logger.info(str(traceback.format_exc()))
                                            print('screenshot:', DirAndTime.get_current_time(), '.png')
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 1)
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, '元素定位超时，请检查元素定位: %s' % str(traceback.format_exc()), 2)
                                        except AssertionError as e:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Failed'
                                            self.error_dic.setdefault(sheet_name, {})[t + 2] = '断言失败'
                                            msg = '步骤"{}"执行失败'.format(step_name)
                                            logger.info(msg)
                                            print(msg)
                                            msg2 = '断言失败: %s' % str(traceback.format_exc())
                                            logger.info(msg2)
                                            print('screenshot:', DirAndTime.get_current_time(), '.png')
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 1)
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg2, 2)
                                        except WebDriverException as e:
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheet_name, {})[t + 2] = '浏览器异常，' \
                                                                                               '请检查浏览器驱动或运行过程中是否被强制关闭'
                                            msg = '步骤"{}"执行失败'.format(step_name)
                                            logger.info(msg)
                                            msg2 = '浏览器异常，请检查浏览器驱动或运行过程中是否被强制关闭'
                                            logger.info(msg2)
                                            print(msg)
                                            print(msg2)
                                            logger.info(str(traceback.format_exc()))
                                            print('screenshot:', DirAndTime.get_current_time(), '.png')
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 1)
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg2, 2)
                                        except Exception as e:
                                            error_info = traceback.format_exc()
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheet_name, {})[t + 2] = error_info
                                            msg = '步骤"{}"执行失败'.format(step_name)
                                            logger.info(msg)
                                            logger.info(error_info)
                                            print(msg)
                                            print('screenshot:', DirAndTime.get_current_time(), '.png')
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 2)
                                        else:
                                            # 写入测试结果
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheet_name, {})[t + 2] = 'Pass'
                                            msg = '步骤"{}"执行成功'.format(step_name)
                                            logger.info(msg)
                                            print(msg)
                                            CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 0)
                                        finally:
                                            # 截图
                                            if self.result_dic.setdefault(sheet_name, {})[t + 2] == 'Skip' or \
                                                    self.result_dic.setdefault(sheet_name, {})[t + 2] == 'Failed':
                                                pic = self.page_action.save_scree_shot(sheet_name, test_case_num)
                                                self.picture_dic.setdefault(sheet_name, {})[t + 2] = pic
                                                CaseInsertLogs.test_case_logs(self.task_obj.task_id, pic, 0)
                                            # 将截图信息以及测试时间存入字典中
                                            now_time = datetime.now()
                                            now_time.strftime('%Y:%m:%d %H:%M:%S')
                                            self.time_dic.setdefault(sheet_name, {})[t + 2] = now_time
                                    else:
                                        continue
                                self.case_num += 1
                                self.page_action.quit_browser()
                                # 写入测试结果
                                for r, v in self.result_dic.items():
                                    for a, b in v.items():
                                        if b == 'Pass':
                                            self.parse_excel.wb[r].cell(int(a), TEST_STEP_RESULT + l).font = Font(
                                                color='33ff33')
                                            self.parse_excel.wb[r].cell(int(a), TEST_STEP_RESULT + l, b)
                                        elif b == 'Failed':
                                            self.parse_excel.wb[r].cell(int(a), TEST_STEP_RESULT + l).font = Font(
                                                color='cc0000')
                                            self.parse_excel.wb[r].cell(int(a), TEST_STEP_RESULT + l, b)
                                        elif b == 'Skip':
                                            self.parse_excel.wb[r].cell(int(a), TEST_STEP_RESULT + l).font = Font(
                                                color='D1D1D1')
                                            self.parse_excel.wb[r].cell(int(a), TEST_STEP_RESULT + l, b)
                                        else:
                                            continue
                                # 通过循环对应 用例编号的步骤的结果，全部为pass的则写入用例Pass，有一条失败的则写入Failed
                                for s, b in enumerate(test_step_num):
                                    # 获取测试结果
                                    if b == test_case_num:
                                        if self.parse_excel.get_cell_value(sheet_name, s + 2, TEST_STEP_RESULT) is None \
                                                or self.parse_excel.get_cell_value(sheet_name, s + 2,
                                                                                   TEST_STEP_RESULT) == '':
                                            continue
                                        elif self.parse_excel.get_cell_value(sheet_name, s + 2,
                                                                             TEST_STEP_RESULT) == 'Pass':
                                            # 将用例测试结果存入字典
                                            self.case_result_dic.setdefault(self.sheet_names[0], {})[index + 2] = 'Pass'
                                        elif self.parse_excel.get_cell_value(sheet_name, s + 2,
                                                                             TEST_STEP_RESULT) == 'Skip':
                                            self.case_result_dic.setdefault(self.sheet_names[0], {})[
                                                index + 2] = 'Failed'
                                            break
                                        else:
                                            self.case_result_dic.setdefault(self.sheet_names[0], {})[
                                                index + 2] = 'Failed'
                                            break
                                now_time = datetime.now()
                                now_time.strftime('%Y:%m:%d %H:%M:%S')
                                # 增加时间写入，以及已运行数量统计
                                self.case_time_dic.setdefault(self.sheet_names[0], {})[index + 2] = now_time
                                # 增加用例时间运行间隔，默认1秒（通过配置文件进行修改）
                                # time.sleep(self.parse_yaml.read_time_wait('case_time'))
                                time.sleep(2)
                            else:
                                continue
                        except Exception as e:
                            now_time = datetime.now()
                            now_time.strftime('%Y:%m:%d %H:%M:%S')
                            self.case_result_dic.setdefault(self.sheet_names[0], {})[index + 2] = 'Failed'
                            self.case_time_dic.setdefault(self.sheet_names[0], {})[index + 2] = now_time
                            self.case_num += 1
                            logger.info(str(traceback.format_exc()))
                    logger.info('正在写入测试结果，请勿关闭界面...')
                    # 读取所有字典，将结果写入excel中
                    for t, v in self.time_dic.items():
                        for a, b in v.items():
                            self.parse_excel.wb[t].cell(int(a), TEST_STEP_END_TIME, b)
                    for e, v in self.error_dic.items():
                        for a, b in v.items():
                            self.parse_excel.wb[e].cell(int(a), TEST_STEP_ERROR, b)
                    for p, v in self.picture_dic.items():
                        for a, b in v.items():
                            self.parse_excel.wb[p].cell(int(a),
                                                        TEST_STEP_PICTURE).value = '=HYPERLINK("{}", "{}")'.format(b, b)
                    for ct, v in self.case_time_dic.items():
                        for a, b in v.items():
                            self.parse_excel.wb[ct].cell(int(a), TEST_CASE_END_TIME, b)
                    for cr, v in self.case_result_dic.items():
                        for a, b in v.items():
                            if b == 'Pass':
                                self.parse_excel.wb[cr].cell(int(a), TEST_CASE_RESULT + l).font = Font(color='33ff33')
                                self.parse_excel.wb[cr].cell(int(a), TEST_CASE_RESULT + l, b)
                            elif b == 'Failed':
                                self.parse_excel.wb[cr].cell(int(a), TEST_CASE_RESULT + l).font = Font(color='cc0000')
                                self.parse_excel.wb[cr].cell(int(a), TEST_CASE_RESULT + l, b)
                            else:
                                continue
                    # 获取excel中'用例工作表'列的不为None的总行数
                    total_case = list(
                        filter(None, self.parse_excel.get_column_value(self.sheet_names[0], TEST_CASE_SHEET)))
                    # 写入excel表的总用例数单元格中
                    self.parse_excel.write_cell_value(self.sheet_names[0], 1, 2, len(total_case) - 1)
                    # 循环执行结果列中为pass的列
                    pass_case = []
                    faild_case = []
                    for pi in list(
                            filter(None, self.parse_excel.get_column_value(self.sheet_names[0], TEST_CASE_RESULT))):
                        if pi.lower() == 'pass':
                            pass_case.append(pi)
                        elif pi.lower() == 'failed':
                            faild_case.append(pi)
                        else:
                            continue
                    # 写入excel表中的通过用例数单元格中
                    self.parse_excel.write_cell_value(self.sheet_names[0], 1, 4, len(pass_case))
                    # 写入excel表中的失败用例数单元格中
                    self.parse_excel.write_cell_value(self.sheet_names[0], 1, 6, len(faild_case))
                    # 循环是否执行列的中n的数量
                    n_case = []
                    for ni in list(
                            filter(None,
                                   self.parse_excel.get_column_value(self.sheet_names[0], TEST_CASE_IS_IMPL_EMENT))):
                        if ni.lower() == 'n':
                            n_case.append(ni)
                    # 写入excel表中的未测试用例的单元格中
                    self.parse_excel.write_cell_value(self.sheet_names[0], 1, 8, len(n_case))
                except Exception as e:
                    logger.info(str(traceback.format_exc()))
                    self.page_action.quit_browser()
                finally:
                    self.parse_excel.wb.save(self.test_data_path)
                    msg = '用例测试结束'
                    logger.info(msg)
                    print(msg)
                    CaseInsertLogs.test_case_logs(self.task_obj.task_id, msg, 3)
                    logger.info('任务：%s, 结束了, 结束时间：%s' %(self.data.get('task_id'), DirAndTime.get_current_time()))
                    current_time = datetime.now()
                    current_time = current_time.strftime('%Y-%m-%d %H:%M:%S')
                    task_obj = TestTask.objects.get(task_id=self.data.get('task_id'))
                    task_obj.task_run_done_time = current_time
                    task_obj.save()

        except Exception as e:
            print('异常结束时，关闭文件流', traceback.format_exc())
            print(e)
            # 异常结束时，关闭文件流
            self.parse_excel.wb.close()
            # 释放任务锁
            conn_redis.delete(self.LOCK_CASE_ID)

    def run_report(self):
        report_path = REPORT_PATH  # 报告存放位置
        time_str = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        file_name = report_path + '\\' + '测试报告' + time_str + '.html'
        fp = open(file_name, 'wb')

        suites = unittest.TestSuite()
        suites.addTest(RunnerTestCase('test_case', self.data))
        runner = HTMLTestRunner(
            title='自动化测试报告',
            description='',
            stream=fp,
            verbosity=2,
        )
        runner.run(suites)
        fp.close()

    def runner_beautiful_report(self):
        suite = unittest.TestSuite()
        suite.addTest(RunnerTestCase('test_case', self.data))
        runner = bf(suite)
        report_path = REPORT_PATH  # 报告存放位置
        time_str = time.strftime('%Y%m%d%H%M', time.localtime(time.time()))
        runner.report(filename='\\' + '测试报告' + time_str + '.html', description=u"测试报告", report_dir=report_path,
                      theme="theme_cyan")
        print('runner.report' + time_str)
